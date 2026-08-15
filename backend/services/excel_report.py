"""Styled .xlsx writer — the house style used by the Brand Portfolio workbooks.

Ported deliberately from the `styled-excel-report` skill's build_workbook.py so
the files this app streams are indistinguishable from the ones built by hand:
same slate palette, same layout, same number formats. The skill's script is a
local tool and cannot be imported from here, so the STYLE CONTRACT is duplicated
rather than shared — if it changes there, change it here too.

Kept spec-driven rather than hand-rolled per report: a sheet is
{name, title, note?, columns[], rows[]}, and a column is
{field, header, format, low_threshold, high_threshold, invert, width}.

Layout per sheet: A2 title -> optional note -> optional summary block -> blank
-> header -> data, frozen below the header with an autofilter over the table.
"""
from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── style contract — see module docstring before touching ────────────────────
SLATE_900 = "334155"   # titles, first column, emphasised values
SLATE_500 = "64748B"   # header text, summary labels, notes
SLATE_600 = "475569"   # body cells
HEADER_FILL = PatternFill("solid", fgColor="F1F5F9")
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ORANGE = "EB6200"      # below low_threshold
GREEN = "15803D"       # at/above high_threshold

DEFAULT_WIDTHS = {"text": 22, "number": 16, "percent": 15, "percent2": 15, "date": 12}
NUM_FORMATS = {"number": "#,##0", "percent": "0.0%", "percent2": "0.00%", "date": "yyyy-mm-dd"}


def _to_number(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) and "." not in str(v) and "e" not in str(v).lower() else f
    except (TypeError, ValueError):
        return None


def _cell_value(row: dict, col: dict):
    v = row.get(col["field"])
    if col.get("format") in ("number", "percent", "percent2"):
        n = _to_number(v)
        return v if n is None else n
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return v


def _build_sheet(ws, spec: dict) -> int:
    rows = spec.get("rows") or []
    columns = spec["columns"]
    for col in columns:
        col.setdefault("header", (col.get("field") or "").upper().replace("_", " "))
        col.setdefault("format", "text")

    ws.cell(row=2, column=1, value=spec.get("title", spec["name"])).font = Font(
        bold=True, size=15, color=SLATE_900
    )

    r = 4
    if spec.get("note"):
        ws.cell(row=r, column=1, value=spec["note"]).font = Font(
            italic=True, size=9, color=SLATE_500
        )
        r += 2

    header_row = r
    for j, col in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=j, value=col["header"])
        c.font = Font(bold=True, size=8, color=SLATE_500)
        c.fill = HEADER_FILL

    zebra = spec.get("zebra", True)
    bold_first = spec.get("bold_first_col", True)
    for i, row in enumerate(rows):
        rr = header_row + 1 + i
        stripe = zebra and i % 2 == 1
        for j, col in enumerate(columns, start=1):
            v = _cell_value(row, col)
            c = ws.cell(row=rr, column=j, value=v)
            if stripe:
                c.fill = ZEBRA_FILL
            fmt = col["format"]
            if j == 1 and bold_first:
                c.font = Font(bold=True, size=9.5, color=SLATE_900)
            elif fmt in ("percent", "percent2", "number") and isinstance(v, (int, float)):
                color, bold = SLATE_600, False
                low, high = col.get("low_threshold"), col.get("high_threshold")
                if col.get("invert"):        # bad-when-high metric
                    if high is not None and v >= high:
                        color, bold = ORANGE, True
                    elif low is not None and v < low:
                        color, bold = GREEN, True
                elif low is not None and v < low:
                    color, bold = ORANGE, True
                elif high is not None and v >= high:
                    color, bold = GREEN, True
                if color == SLATE_600 and fmt in ("percent", "percent2"):
                    color, bold = SLATE_900, True
                c.font = Font(bold=bold, size=9, color=color)
            else:
                c.font = Font(size=9, color=SLATE_600)
            if fmt in NUM_FORMATS:
                c.number_format = NUM_FORMATS[fmt]

    for j, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = col.get(
            "width", DEFAULT_WIDTHS[col["format"]]
        )

    ws.freeze_panes = f"A{header_row + 1}"
    if columns:
        last = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A{header_row}:{last}{header_row + len(rows)}"
    return len(rows)


def build_workbook(sheets: list[dict]) -> BytesIO:
    """Render sheet specs to an in-memory .xlsx. Empty sheets are kept, so a
    competitor with no comp-only products still gets the tab rather than a file
    whose shape changes with the data."""
    wb = openpyxl.Workbook()
    used: set[str] = set()
    for i, spec in enumerate(sheets):
        name = str(spec.get("name") or "Sheet")
        for ch in ':\\/?*[]':
            name = name.replace(ch, " ")
        name = name.strip()[:31] or "Sheet"
        base, n = name, 2
        while name in used:
            suffix = f" ({n})"
            name = base[:31 - len(suffix)] + suffix
            n += 1
        used.add(name)
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = name
        _build_sheet(ws, spec)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
